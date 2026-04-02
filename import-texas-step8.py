#!/usr/bin/env python3
"""
Step 8: Import Texas HHSC ALF + DAHS + ALF Closures into Supabase facilities table.

This script:
1. Loads all existing TX records from the facilities table
2. Processes ALF (Assisted Living Facility) directory — merges with existing, inserts new
3. Parses ALF Closures PDF — marks matched existing records inactive
4. Processes DAHS (Day Activity and Health Services) directory — inserts as adult_day
5. Reports match/insert/update statistics

Data sources:
- ALF: AL.xlsx (Texas HHSC, 2,006 active facilities as of 03/26/2026)
- Closures: alf_closures.pdf (Texas HHSC, ~792 closed facilities)
- DAHS: DAHS.xlsx (Texas HHSC, 398 active facilities as of 03/26/2026)

Run from project root (our-turn-to-care/):
    caffeinate -i python3 import-texas-step8.py

Requirements: supabase-py, openpyxl, pdfplumber (install with pip3 if needed)
    pip3 install openpyxl pdfplumber
"""

import os
import re
import json
import sys
from datetime import datetime

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# File paths — adjust if your files are in a different location
ALF_FILE = "AL.xlsx"
CLOSURES_FILE = "alf_closures.pdf"
DAHS_FILE = "DAHS.xlsx"

# Batch size for Supabase upserts (keep small to avoid free-tier timeouts)
BATCH_SIZE = 200

# Matching thresholds
NAME_MATCH_THRESHOLD = 0.65  # 65% bigram similarity to consider a name match


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_env():
    """Read Supabase credentials from .env.local"""
    env = {}
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env.local')
    if not os.path.exists(env_path):
        print(f"ERROR: .env.local not found at {env_path}")
        sys.exit(1)
    with open(env_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, val = line.split('=', 1)
                env[key.strip()] = val.strip()
    return env


def normalize_address(addr):
    """
    Normalize an address for matching:
    - Uppercase, remove punctuation, normalize abbreviations, collapse whitespace
    """
    if not addr:
        return ""
    s = addr.upper().strip()
    s = re.sub(r'[.,#]', '', s)
    replacements = {
        ' STREET': ' ST',
        ' AVENUE': ' AVE',
        ' BOULEVARD': ' BLVD',
        ' DRIVE': ' DR',
        ' ROAD': ' RD',
        ' LANE': ' LN',
        ' COURT': ' CT',
        ' PLACE': ' PL',
        ' CIRCLE': ' CIR',
        ' HIGHWAY': ' HWY',
        ' PARKWAY': ' PKWY',
        ' NORTH': ' N',
        ' SOUTH': ' S',
        ' EAST': ' E',
        ' WEST': ' W',
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def normalize_zip(z):
    """Extract 5-digit zip code."""
    if not z:
        return ""
    z = str(z).strip()
    match = re.match(r'(\d{5})', z)
    return match.group(1) if match else z


def simple_similarity(a, b):
    """Bigram (Dice coefficient) similarity between two strings. Returns 0.0-1.0."""
    if not a or not b:
        return 0.0
    a = a.upper().strip()
    b = b.upper().strip()
    if a == b:
        return 1.0
    def bigrams(s):
        return set(s[i:i+2] for i in range(len(s)-1))
    ba = bigrams(a)
    bb = bigrams(b)
    if not ba or not bb:
        return 0.0
    overlap = len(ba & bb)
    return (2.0 * overlap) / (len(ba) + len(bb))


def clean_phone(phone):
    """Normalize phone number to (XXX) XXX-XXXX format."""
    if not phone:
        return None
    digits = re.sub(r'\D', '', str(phone))
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    elif len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return str(phone).strip() if str(phone).strip() else None


def safe_int(val):
    """Convert to int safely, return None if not possible."""
    if val is None or val == '' or val == 0:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def parse_geo_location(geo_str):
    """Parse 'lat,lon' string from Texas Geo Location column. Returns (lat, lon) or (None, None)."""
    if not geo_str or not str(geo_str).strip():
        return None, None
    try:
        parts = str(geo_str).split(',')
        if len(parts) == 2:
            lat = float(parts[0].strip())
            lon = float(parts[1].strip())
            # Sanity check — Texas coords should be roughly lat 25-37, lon -107 to -93
            if 25 <= lat <= 37 and -107 <= lon <= -93:
                return lat, lon
            # Maybe reversed?
            if 25 <= lon <= 37 and -107 <= lat <= -93:
                return lon, lat
    except (ValueError, TypeError):
        pass
    return None, None


def format_date(dt_val):
    """Convert datetime object or string to ISO date string."""
    if not dt_val:
        return None
    if isinstance(dt_val, datetime):
        return dt_val.strftime('%Y-%m-%d')
    if isinstance(dt_val, str):
        for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
            try:
                return datetime.strptime(dt_val.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def read_excel_rows(filepath, header_row=2, data_start_row=3):
    """
    Read an Excel file and return list of dicts.
    header_row: 1-based row number for column headers
    data_start_row: 1-based row number where data begins
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Read headers
    headers = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(row)]
        break

    # Read data
    records = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row[0] is None:
            continue
        rec = {}
        for i, val in enumerate(row):
            if i < len(headers):
                rec[headers[i]] = val
        records.append(rec)

    wb.close()
    return headers, records


def parse_closures_pdf(filepath):
    """
    Parse the ALF Closures PDF and extract structured closure records.
    Each entry has: Closure Date, Facility ID, Owner, Facility Name, Address, City, State, Zip, County, Service Type.
    """
    import pdfplumber

    closures = []
    pdf = pdfplumber.open(filepath)

    for page_num, page in enumerate(pdf.pages):
        text = page.extract_text()
        if not text:
            continue

        # Split into lines
        lines = text.split('\n')

        # Parse entries — each starts with "Closure Date"
        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if line.startswith('Closure Date'):
                entry = {}

                # Line: "Closure Date MM/DD/YYYY Facility ID XXXXXX"
                m = re.search(r'Closure Date\s+(\d{1,2}/\d{1,2}/\d{4})\s+Facility ID\s+(\d+)', line)
                if m:
                    entry['closure_date'] = m.group(1)
                    entry['facility_id'] = m.group(2)
                else:
                    i += 1
                    continue

                # Next lines: Owner/Operator, Facility, Address, City/State/Zip, County, Service Type
                # Consume up to 8 more lines looking for these fields
                j = i + 1
                raw_lines = []
                while j < len(lines) and j <= i + 10:
                    next_line = lines[j].strip()
                    if next_line.startswith('Closure Date') or next_line.startswith('Texas Health and Human Services'):
                        break
                    raw_lines.append(next_line)
                    j += 1

                # Parse the raw lines
                for rl in raw_lines:
                    if rl.startswith('Owner/Operator'):
                        entry['owner'] = rl.replace('Owner/Operator', '').strip()
                    elif rl.startswith('Facility') and 'Facility ID' not in rl:
                        entry['facility_name'] = rl.replace('Facility', '', 1).strip()
                    elif rl.startswith('Address'):
                        # "Address 1237 Cedarbrook Trail Bed Designations"
                        addr = rl.replace('Address', '', 1).strip()
                        addr = re.sub(r'\s*Bed Designations.*', '', addr).strip()
                        entry['address'] = addr
                    elif rl.startswith('County'):
                        # "County Dallas Total Capacity 0"
                        county = rl.replace('County', '', 1).strip()
                        county = re.sub(r'\s*Total Capacity.*', '', county).strip()
                        entry['county'] = county
                    elif rl.startswith('Service Type'):
                        entry['service_type'] = rl.replace('Service Type', '').strip()
                    elif 'Private Pay Beds' in rl:
                        # This line often has "City TX ZIPCODE Private Pay Beds X"
                        # or just "City TX ZIPCODE"
                        city_line = re.sub(r'\s*Private Pay Beds.*', '', rl).strip()
                        # Try to extract city, state, zip
                        m2 = re.match(r'(.+?)\s+(TX|Texas)\s+(\d{5}(?:-\d{4})?)', city_line, re.IGNORECASE)
                        if m2:
                            entry['city'] = m2.group(1).strip()
                            entry['state'] = 'TX'
                            entry['zip'] = m2.group(3).strip()

                if entry.get('facility_id') and (entry.get('facility_name') or entry.get('address')):
                    closures.append(entry)

                i = j  # Skip past the lines we consumed
            else:
                i += 1

    pdf.close()
    return closures


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("Step 8: Texas HHSC ALF + DAHS + Closures Import")
    print("=" * 70)
    print()

    # Check files exist
    for f in [ALF_FILE, DAHS_FILE, CLOSURES_FILE]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            print(f"  Expected in: {os.getcwd()}")
            print(f"  Make sure {f} is in the same directory as this script.")
            sys.exit(1)

    # Load Supabase credentials
    env = load_env()
    url = env.get('NEXT_PUBLIC_SUPABASE_URL')
    key = env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        print("ERROR: Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")
        sys.exit(1)

    from supabase import create_client
    supabase = create_client(url, key)
    print(f"Connected to Supabase: {url}")
    print()

    # ==================================================================
    # PHASE 1: Load existing TX records from Supabase
    # ==================================================================
    print("-" * 70)
    print("PHASE 1: Loading existing Texas facilities from Supabase...")
    print("-" * 70)

    existing_tx = []
    page_size = 1000
    offset = 0
    while True:
        result = supabase.table('facilities').select(
            'facility_id, source, source_id, facility_name, address_line_1, city, state, zip_code, '
            'facility_type, latitude, longitude, is_active, license_number, accepts_medicaid, bed_count'
        ).eq('state', 'TX').range(offset, offset + page_size - 1).execute()

        batch = result.data
        if not batch:
            break
        existing_tx.extend(batch)
        offset += page_size
        if len(batch) < page_size:
            break

    print(f"  Loaded {len(existing_tx)} existing TX records from Supabase")

    # Build lookup indexes
    zip_addr_index = {}
    city_name_index = {}

    for rec in existing_tx:
        z = normalize_zip(rec.get('zip_code', ''))
        a = normalize_address(rec.get('address_line_1', ''))
        if z and a:
            key_za = f"{z}|{a}"
            if key_za not in zip_addr_index:
                zip_addr_index[key_za] = []
            zip_addr_index[key_za].append(rec)

        city = (rec.get('city') or '').upper().strip()
        name = (rec.get('facility_name') or '').upper().strip()
        if city and name:
            key_cn = f"{city}|{name}"
            if key_cn not in city_name_index:
                city_name_index[key_cn] = []
            city_name_index[key_cn].append(rec)

    print(f"  Built zip+address index: {len(zip_addr_index)} unique entries")
    print(f"  Built city+name index: {len(city_name_index)} unique entries")
    print()

    def find_match(address, zip_code, name, city):
        """Try to find a matching existing record using zip+address then fuzzy name."""
        z = normalize_zip(zip_code)
        a = normalize_address(address)

        # Strategy 1: Exact zip + normalized address match
        if z and a:
            key_za = f"{z}|{a}"
            if key_za in zip_addr_index:
                return zip_addr_index[key_za][0]

        # Strategy 2: Same city + fuzzy name match
        c = city.upper().strip() if city else ''
        n = name.upper().strip() if name else ''
        if c and n:
            best_match = None
            best_score = 0
            for key_cn, recs in city_name_index.items():
                if key_cn.startswith(c + '|'):
                    existing_name = key_cn.split('|', 1)[1]
                    score = simple_similarity(n, existing_name)
                    if score > best_score:
                        best_score = score
                        best_match = recs[0]

            if best_match and best_score >= NAME_MATCH_THRESHOLD:
                match_zip = normalize_zip(best_match.get('zip_code', ''))
                if not z or not match_zip or z == match_zip:
                    return best_match

        return None

    # ==================================================================
    # PHASE 2: Process ALF (Assisted Living Facility) directory
    # ==================================================================
    print("-" * 70)
    print("PHASE 2: Processing ALF directory (AL.xlsx)...")
    print("-" * 70)

    headers, alf_records = read_excel_rows(ALF_FILE, header_row=2, data_start_row=3)
    print(f"  Read {len(alf_records)} ALF records from {ALF_FILE}")
    print(f"  Columns: {len(headers)}")
    print()

    alf_updates = []      # Records to UPDATE (matched existing)
    alf_inserts = []      # Records to INSERT (new facilities)
    alf_matched_ids = set()  # Track which existing records we've matched

    print("  Matching ALF records against existing database...")
    for i, row in enumerate(alf_records):
        address = str(row.get('Physical Address', '') or '').strip()
        city = str(row.get('Physical Address CITY', '') or '').strip()
        zip_code = str(row.get('Physical Address Zipcode', '') or '').strip()
        fac_name = str(row.get('Facility Name', '') or '').strip()
        facility_id_tx = str(row.get('Facility ID', '') or '').strip()

        # Parse geo coordinates
        lat, lon = parse_geo_location(row.get('Geo Location', ''))

        # Parse key fields
        license_no = str(row.get('License No', '') or '').strip()
        service_type = str(row.get('Service  Type', '') or '').strip()  # Note: double space in column name
        total_capacity = safe_int(row.get('Total Licensed Capacity'))
        alz_capacity = safe_int(row.get('Alzheimer Capacity'))
        alz_cert_no = str(row.get('Alzheimer Certificate No', '') or '').strip() or None
        alz_cert_eff = format_date(row.get('Alzheimer Certificate Effective Date'))
        alz_cert_exp = format_date(row.get('Alzheimer Expiration Date'))
        owner = str(row.get('Owner_', '') or '').strip()
        entity_type = str(row.get('Type of Entity', '') or '').strip()
        phone = clean_phone(row.get('Facility Phone Number'))
        email = str(row.get('Provider Email', '') or '').strip() or None
        administrator = str(row.get('Administrator', '') or '').strip() or None
        mgmt_company = str(row.get('Management Company_', '') or '').strip() or None
        county = str(row.get('County', '') or '').strip() or None
        lic_eff_date = format_date(row.get('License Effective Date'))
        lic_exp_date = format_date(row.get('License Expiration Date'))
        initial_lic_date = format_date(row.get('Initial License Date'))
        medicaid_provider_num = str(row.get('Medicaid Provider Number', '') or '').strip() or None
        medicaid_beds = safe_int(row.get('Medicaid Only Beds'))
        medicare_medicaid_beds = safe_int(row.get('Medicaid / Medicare Beds'))

        # Determine accepts_medicaid flag
        has_medicaid = bool(medicaid_provider_num) or (medicaid_beds and medicaid_beds > 0) or (medicare_medicaid_beds and medicare_medicaid_beds > 0)

        # Build raw_data JSONB with all Texas-specific fields
        raw_data = {
            'tx_hhsc': {
                'facility_id': facility_id_tx,
                'facility_name': fac_name,
                'program_type': str(row.get('Program Type', '') or '').strip(),
                'service_type': service_type,
                'license_no': license_no,
                'license_effective_date': lic_eff_date,
                'license_expiration_date': lic_exp_date,
                'initial_license_date': initial_lic_date,
                'total_licensed_capacity': total_capacity,
                'licensed_only_beds': safe_int(row.get('Licensed Only Beds')),
                'medicaid_only_beds': medicaid_beds,
                'medicare_only_beds': safe_int(row.get('Medicare Only Beds')),
                'medicaid_medicare_beds': medicare_medicaid_beds,
                'alzheimer_capacity': alz_capacity,
                'alzheimer_certificate_no': alz_cert_no,
                'alzheimer_cert_effective': alz_cert_eff,
                'alzheimer_cert_expiration': alz_cert_exp,
                'owner': owner,
                'entity_type': entity_type,
                'administrator': administrator,
                'management_company': mgmt_company,
                'state_region': str(row.get('State Region', '') or '').strip(),
                'hhsc_suboffice': str(row.get('HHSC SubOffice', '') or '').strip(),
                'facility_certified': str(row.get('Facility Certified', '') or '').strip(),
            }
        }

        match = find_match(address, zip_code, fac_name, city)

        if match:
            fid = match['facility_id']
            if fid not in alf_matched_ids:
                alf_matched_ids.add(fid)
                update = {
                    'facility_id': fid,
                    'license_number': license_no if license_no else facility_id_tx,
                    'license_status': 'LICENSED',
                    'is_active': True,
                    'raw_data': json.dumps(raw_data),
                }
                # Update bed_count if Texas has a value
                if total_capacity and total_capacity > 0:
                    update['bed_count'] = total_capacity
                # Add legal_name (owner entity)
                if owner:
                    update['legal_name'] = owner
                # Add phone if available
                if phone:
                    update['phone'] = phone
                # Add email if available
                if email:
                    update['email'] = email
                # Add county if available
                if county:
                    update['county'] = county
                # Add coordinates if existing record is missing them
                if lat and lon and (not match.get('latitude') or not match.get('longitude')):
                    update['latitude'] = lat
                    update['longitude'] = lon
                # Flag Medicaid acceptance
                if has_medicaid:
                    update['accepts_medicaid'] = True
                # Update ownership_type from entity_type
                if entity_type:
                    update['ownership_type'] = entity_type

                alf_updates.append(update)
        else:
            # New facility — prepare INSERT
            record = {
                'source': 'tx_hhsc',
                'source_id': facility_id_tx,
                'facility_name': fac_name,
                'legal_name': owner if owner else None,
                'facility_type': 'assisted_living',
                'ownership_type': entity_type if entity_type else None,
                'address_line_1': address or 'Unknown',
                'city': city or 'Unknown',
                'state': 'TX',
                'zip_code': normalize_zip(zip_code) or '00000',
                'county': county,
                'phone': phone,
                'email': email,
                'bed_count': total_capacity,
                'license_number': license_no if license_no else facility_id_tx,
                'license_status': 'LICENSED',
                'latitude': lat,
                'longitude': lon,
                'accepts_medicaid': has_medicaid if has_medicaid else None,
                'is_active': True,
                'raw_data': json.dumps(raw_data),
            }
            alf_inserts.append(record)

        if (i + 1) % 500 == 0:
            print(f"    Processed {i + 1}/{len(alf_records)} ALF records...")

    # Count Alzheimer-capable facilities
    alz_count = sum(1 for r in alf_records if safe_int(r.get('Alzheimer Capacity')) and safe_int(r.get('Alzheimer Capacity')) > 0)

    print(f"\n  ALF matching results:")
    print(f"    Matched existing records (will UPDATE): {len(alf_updates)}")
    print(f"    New facilities (will INSERT): {len(alf_inserts)}")
    print(f"    Facilities with Alzheimer capacity: {alz_count} (stored in raw_data)")
    print()

    # Execute ALF updates
    print("  Executing ALF updates...")
    update_count = 0
    for i in range(0, len(alf_updates), BATCH_SIZE):
        batch = alf_updates[i:i+BATCH_SIZE]
        for rec in batch:
            fid = rec.pop('facility_id')
            try:
                supabase.table('facilities').update(rec).eq('facility_id', fid).execute()
                update_count += 1
            except Exception as e:
                print(f"    WARNING: Update failed for {fid}: {e}")
        print(f"    Updated {min(i + BATCH_SIZE, len(alf_updates))}/{len(alf_updates)}")
    print(f"  ALF updates complete: {update_count} records enriched")
    print()

    # Execute ALF inserts (deduplicate first)
    print("  Executing ALF inserts...")
    insert_count = 0
    seen_source_ids = set()
    deduped_inserts = []
    for rec in alf_inserts:
        sid = rec['source_id']
        if sid not in seen_source_ids:
            seen_source_ids.add(sid)
            deduped_inserts.append(rec)
    if len(deduped_inserts) < len(alf_inserts):
        print(f"    Deduplicated: {len(alf_inserts)} -> {len(deduped_inserts)} (removed {len(alf_inserts) - len(deduped_inserts)} duplicates)")

    for i in range(0, len(deduped_inserts), BATCH_SIZE):
        batch = deduped_inserts[i:i+BATCH_SIZE]
        try:
            supabase.table('facilities').upsert(batch, on_conflict='source,source_id').execute()
            insert_count += len(batch)
        except Exception as e:
            print(f"    WARNING: Insert batch failed at offset {i}: {e}")
            for rec in batch:
                try:
                    supabase.table('facilities').upsert([rec], on_conflict='source,source_id').execute()
                    insert_count += 1
                except Exception as e2:
                    print(f"      SKIP: {rec.get('facility_name', '?')}: {e2}")
        print(f"    Inserted {min(i + BATCH_SIZE, len(deduped_inserts))}/{len(deduped_inserts)}")
    print(f"  ALF inserts complete: {insert_count} new records added")
    print()

    # ==================================================================
    # PHASE 3: Process ALF Closures PDF
    # ==================================================================
    print("-" * 70)
    print("PHASE 3: Processing ALF Closures (alf_closures.pdf)...")
    print("-" * 70)

    # Reload TX records to include newly inserted ALF records
    print("  Reloading TX records from Supabase (includes new ALF inserts)...")
    existing_tx_2 = []
    offset = 0
    while True:
        result = supabase.table('facilities').select(
            'facility_id, source, source_id, facility_name, address_line_1, city, state, zip_code, '
            'facility_type, latitude, longitude, is_active, license_number, accepts_medicaid'
        ).eq('state', 'TX').range(offset, offset + page_size - 1).execute()
        batch = result.data
        if not batch:
            break
        existing_tx_2.extend(batch)
        offset += page_size
        if len(batch) < page_size:
            break
    print(f"  Loaded {len(existing_tx_2)} TX records (post-ALF import)")

    # Rebuild indexes
    zip_addr_index_2 = {}
    city_name_index_2 = {}
    # Also build a source_id index for tx_hhsc records (to match by Facility ID)
    hhsc_id_index = {}

    for rec in existing_tx_2:
        z = normalize_zip(rec.get('zip_code', ''))
        a = normalize_address(rec.get('address_line_1', ''))
        if z and a:
            key_za = f"{z}|{a}"
            if key_za not in zip_addr_index_2:
                zip_addr_index_2[key_za] = []
            zip_addr_index_2[key_za].append(rec)

        city = (rec.get('city') or '').upper().strip()
        name = (rec.get('facility_name') or '').upper().strip()
        if city and name:
            key_cn = f"{city}|{name}"
            if key_cn not in city_name_index_2:
                city_name_index_2[key_cn] = []
            city_name_index_2[key_cn].append(rec)

        # Index tx_hhsc records by their HHSC Facility ID (source_id)
        if rec.get('source') == 'tx_hhsc':
            hhsc_id_index[rec['source_id']] = rec

    def find_match_v2(address, zip_code, name, city, hhsc_facility_id=None):
        """Match with updated indexes. Also tries HHSC Facility ID for tx_hhsc records."""
        # Strategy 0: Direct HHSC Facility ID match (for closures)
        if hhsc_facility_id and hhsc_facility_id in hhsc_id_index:
            return hhsc_id_index[hhsc_facility_id]

        z = normalize_zip(zip_code)
        a = normalize_address(address)

        if z and a:
            key_za = f"{z}|{a}"
            if key_za in zip_addr_index_2:
                return zip_addr_index_2[key_za][0]

        c = city.upper().strip() if city else ''
        n = name.upper().strip() if name else ''
        if c and n:
            best_match = None
            best_score = 0
            for key_cn, recs in city_name_index_2.items():
                if key_cn.startswith(c + '|'):
                    existing_name = key_cn.split('|', 1)[1]
                    score = simple_similarity(n, existing_name)
                    if score > best_score:
                        best_score = score
                        best_match = recs[0]

            if best_match and best_score >= NAME_MATCH_THRESHOLD:
                match_zip = normalize_zip(best_match.get('zip_code', ''))
                if not z or not match_zip or z == match_zip:
                    return best_match

        return None

    # Parse closures from PDF
    print("  Parsing closures from PDF...")
    closure_records = parse_closures_pdf(CLOSURES_FILE)
    print(f"  Parsed {len(closure_records)} closure entries from PDF")
    print()

    # Match closures against existing records
    closure_updates = []
    closure_matched_ids = set()
    closure_no_match = 0

    print("  Matching closure records against existing database...")
    for cr in closure_records:
        match = find_match_v2(
            cr.get('address', ''),
            cr.get('zip', ''),
            cr.get('facility_name', ''),
            cr.get('city', ''),
            hhsc_facility_id=cr.get('facility_id')
        )

        if match:
            fid = match['facility_id']
            if fid not in closure_matched_ids and fid not in alf_matched_ids:
                closure_matched_ids.add(fid)
                closure_updates.append({
                    'facility_id': fid,
                    'is_active': False,
                    'license_status': 'CLOSED',
                    'raw_data': json.dumps({
                        'tx_hhsc_closure': {
                            'facility_id': cr.get('facility_id', ''),
                            'facility_name': cr.get('facility_name', ''),
                            'owner': cr.get('owner', ''),
                            'closure_date': cr.get('closure_date', ''),
                            'address': cr.get('address', ''),
                            'city': cr.get('city', ''),
                            'zip': cr.get('zip', ''),
                            'county': cr.get('county', ''),
                            'service_type': cr.get('service_type', ''),
                        }
                    }),
                })
        else:
            closure_no_match += 1

    print(f"  Closure matching results:")
    print(f"    Matched existing records (will mark inactive): {len(closure_updates)}")
    print(f"    No match found (facility not in database): {closure_no_match}")
    print()

    # Execute closure updates
    print("  Executing closure updates...")
    closure_count = 0
    for rec in closure_updates:
        fid = rec.pop('facility_id')
        try:
            supabase.table('facilities').update(rec).eq('facility_id', fid).execute()
            closure_count += 1
        except Exception as e:
            print(f"    WARNING: Closure update failed for {fid}: {e}")
    print(f"  Closures complete: {closure_count} records marked inactive")
    print()

    # ==================================================================
    # PHASE 4: Process DAHS (Day Activity and Health Services) directory
    # ==================================================================
    print("-" * 70)
    print("PHASE 4: Processing DAHS directory (DAHS.xlsx)...")
    print("-" * 70)

    headers_dahs, dahs_records = read_excel_rows(DAHS_FILE, header_row=2, data_start_row=3)
    print(f"  Read {len(dahs_records)} DAHS records from {DAHS_FILE}")
    print()

    dahs_inserts = []
    for row in dahs_records:
        fac_name = str(row.get('Facility Name', '') or '').strip()
        facility_id_tx = str(row.get('Facility ID', '') or '').strip()
        address = str(row.get('Physical Address', '') or '').strip()
        city = str(row.get('Physical Address CITY', '') or '').strip()
        zip_code = str(row.get('Physical Address Zipcode', '') or '').strip()
        county = str(row.get('County', '') or '').strip() or None
        phone = clean_phone(row.get('Facility Phone Number'))
        email = str(row.get('Provider Email', '') or '').strip() or None
        lat, lon = parse_geo_location(row.get('Geo Location', ''))
        total_capacity = safe_int(row.get('Total Licensed Capacity'))
        license_no = str(row.get('License No', '') or '').strip()
        owner = str(row.get('Owner_', '') or '').strip()
        entity_type = str(row.get('Type of Entity', '') or '').strip()
        administrator = str(row.get('Administrator', '') or '').strip() or None
        mgmt_company = str(row.get('Management Company_', '') or '').strip() or None
        program_type = str(row.get('Program Type', '') or '').strip()
        service_type = str(row.get('Service  Type', '') or '').strip()
        medicaid_provider_num = str(row.get('Medicaid Provider Number', '') or '').strip() or None
        medicaid_beds = safe_int(row.get('Medicaid Only Beds'))
        medicare_medicaid_beds = safe_int(row.get('Medicaid / Medicare Beds'))
        has_medicaid = bool(medicaid_provider_num) or (medicaid_beds and medicaid_beds > 0) or (medicare_medicaid_beds and medicare_medicaid_beds > 0)

        raw_data = {
            'tx_dahs': {
                'facility_id': facility_id_tx,
                'facility_name': fac_name,
                'program_type': program_type,
                'service_type': service_type,
                'license_no': license_no,
                'license_effective_date': format_date(row.get('License Effective Date')),
                'license_expiration_date': format_date(row.get('License Expiration Date')),
                'initial_license_date': format_date(row.get('Initial License Date')),
                'total_licensed_capacity': total_capacity,
                'licensed_only_beds': safe_int(row.get('Licensed Only Beds')),
                'medicaid_only_beds': medicaid_beds,
                'medicare_only_beds': safe_int(row.get('Medicare Only Beds')),
                'medicaid_medicare_beds': medicare_medicaid_beds,
                'alzheimer_capacity': safe_int(row.get('Alzheimer Capacity')),
                'owner': owner,
                'entity_type': entity_type,
                'administrator': administrator,
                'management_company': mgmt_company,
                'state_region': str(row.get('State Region', '') or '').strip(),
                'hhsc_suboffice': str(row.get('HHSC SubOffice', '') or '').strip(),
            }
        }

        record = {
            'source': 'tx_dahs',
            'source_id': facility_id_tx,
            'facility_name': fac_name,
            'legal_name': owner if owner else None,
            'facility_type': 'adult_day',
            'ownership_type': entity_type if entity_type else None,
            'address_line_1': address or 'Unknown',
            'city': city or 'Unknown',
            'state': 'TX',
            'zip_code': normalize_zip(zip_code) or '00000',
            'county': county,
            'phone': phone,
            'email': email,
            'bed_count': total_capacity,
            'license_number': license_no if license_no else facility_id_tx,
            'license_status': 'LICENSED',
            'latitude': lat,
            'longitude': lon,
            'accepts_medicaid': has_medicaid if has_medicaid else None,
            'is_active': True,
            'raw_data': json.dumps(raw_data),
        }
        dahs_inserts.append(record)

    # Deduplicate DAHS inserts
    seen_dahs_ids = set()
    deduped_dahs = []
    for rec in dahs_inserts:
        sid = rec['source_id']
        if sid not in seen_dahs_ids:
            seen_dahs_ids.add(sid)
            deduped_dahs.append(rec)
    if len(deduped_dahs) < len(dahs_inserts):
        print(f"  Deduplicated: {len(dahs_inserts)} -> {len(deduped_dahs)}")

    print(f"  DAHS records to insert: {len(deduped_dahs)}")
    print()

    # Execute DAHS inserts
    print("  Executing DAHS inserts...")
    dahs_count = 0
    for i in range(0, len(deduped_dahs), BATCH_SIZE):
        batch = deduped_dahs[i:i+BATCH_SIZE]
        try:
            supabase.table('facilities').upsert(batch, on_conflict='source,source_id').execute()
            dahs_count += len(batch)
        except Exception as e:
            print(f"    WARNING: DAHS insert batch failed at offset {i}: {e}")
            for rec in batch:
                try:
                    supabase.table('facilities').upsert([rec], on_conflict='source,source_id').execute()
                    dahs_count += 1
                except Exception as e2:
                    print(f"      SKIP: {rec.get('facility_name', '?')}: {e2}")
        print(f"    Inserted {min(i + BATCH_SIZE, len(deduped_dahs))}/{len(deduped_dahs)}")
    print(f"  DAHS inserts complete: {dahs_count} new adult_day records added")
    print()

    # ==================================================================
    # PHASE 5: Summary
    # ==================================================================
    print("=" * 70)
    print("STEP 8 COMPLETE — SUMMARY")
    print("=" * 70)
    print()
    print("ALF (Assisted Living Facility) Directory:")
    print(f"  Source records: {len(alf_records)}")
    print(f"  Matched & enriched existing records: {update_count}")
    print(f"  New facilities inserted: {insert_count}")
    print(f"  Facilities with Alzheimer capacity: {alz_count} (in raw_data)")
    print()
    print("ALF Closures (from PDF):")
    print(f"  Closure entries parsed from PDF: {len(closure_records)}")
    print(f"  Matched & marked inactive: {closure_count}")
    print(f"  No match found: {closure_no_match}")
    print()
    print("DAHS (Day Activity and Health Services):")
    print(f"  Source records: {len(dahs_records)}")
    print(f"  New adult_day facilities inserted: {dahs_count}")
    print()

    # Count new records without coordinates
    no_coords_alf = sum(1 for r in deduped_inserts if not r.get('latitude') or not r.get('longitude'))
    no_coords_dahs = sum(1 for r in deduped_dahs if not r.get('latitude') or not r.get('longitude'))
    print(f"Records missing lat/lon: {no_coords_alf} ALF + {no_coords_dahs} DAHS = {no_coords_alf + no_coords_dahs} total")
    if no_coords_alf + no_coords_dahs > 0:
        print("  These can be geocoded using the Census Geocoder (geocode-facilities.py)")
    print()

    # Final counts
    try:
        count_result = supabase.table('facilities').select('facility_id', count='exact').execute()
        total = count_result.count if hasattr(count_result, 'count') and count_result.count else '?'
        tx_result = supabase.table('facilities').select('facility_id', count='exact').eq('state', 'TX').execute()
        tx_total = tx_result.count if hasattr(tx_result, 'count') and tx_result.count else '?'
        print(f"Total facilities in database: {total}")
        print(f"Texas facilities: {tx_total}")
    except Exception as e:
        print(f"  Could not query final counts: {e}")
    print()


if __name__ == '__main__':
    main()